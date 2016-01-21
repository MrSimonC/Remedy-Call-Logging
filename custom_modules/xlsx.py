import csv
import warnings
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Style
from collections import OrderedDict
__version__ = 1.1
warnings.filterwarnings('ignore')  # Turn off openpyxl Discarded range warnings


class XlsxTools:
    def create_document(self, array_of_dict, tab_name, xlsx_filename):
        """
        Write an xlsx file from [({'header1': 'dataA', 'header2':'dataB'}), ({'header1': 'dataC', 'header2':'dataD'})]
        """
        self.xlsx_filename = xlsx_filename
        self.wb = Workbook()
        self.ws = self.wb.active    #default: "Sheet"
        self.ws.title = tab_name
        self.write_data(array_of_dict)
        self.format(array_of_dict)
        self.wb.save(xlsx_filename)

    def write_data(self, array_of_dict):
        self.ws.append(list(array_of_dict[0].keys()))  # headers
        for row in array_of_dict:  # data
            self.ws.append(['%s' % data for data in list(row.values())])

    def add_work_sheet(self, array_of_dict, tab_name, xlsx_filename=""):
        """
        Add worksheet. If xlsx_filename, open. If not, assume self.wb is loaded
        """
        if xlsx_filename:
            self.xlsx_filename = xlsx_filename
            self.wb = load_workbook(self.xlsx_filename)
        else:
            try:
                self.wb    #createDocument creates self.wb and should have been called
            except AttributeError:
                raise AttributeError("addWorkSheet: No filename provided, or didn't call createDocument first")
        self.ws = self.wb.create_sheet()
        self.ws.title = tab_name
        self.write_data(array_of_dict)
        self.format(array_of_dict)
        self.wb.save(self.xlsx_filename)

    def format(self, array_of_dict):
        self.freeze_panes_first_row()
        self.auto_fit(array_of_dict)
        self.auto_filter()
        self.make_first_row_bold()

    def auto_fit(self, array_of_dict):
        """
        AutoFit Column Widths
        Sets the column widths automatically based on length. Assumes you've created self.wb / self.ws
        from http://stackoverflow.com/questions/13197574/python-openpyxl-column-width-size-adjust
        Ignore header length, find the max string length of each cell contents
        """
        column_widths = []
        for header_name in list(array_of_dict[0].keys()):
            # str() is needed in len(str()) as len(int) is an error
            column_widths.append(max(len(str(data[header_name])) if data[header_name] != None else 0 for data in array_of_dict))

        for i, column_width in enumerate(column_widths):
            if column_width < 10:
                column_width = 10
            if column_width > 50:
                column_width = 50
            self.ws.column_dimensions[get_column_letter(i+1)].width = column_width

    def header_row_reference(self):
        return "A1:" + get_column_letter(len(self.ws.rows[0])) + "1"

    def auto_filter(self):
        self.ws.auto_filter.ref = self.header_row_reference()

    def freeze_panes_first_row_all(self):
        #need openpyxl v2.2b1 = pip install -U --pre openpyxl, to fix load_workbook ignoring freeze_panes settings
        # 16/Jun/15 = v.2.2.3 official pip insstall
        #https://bitbucket.org/openpyxl/openpyxl/issue/427/wsfreeze_panes-not-retained-when-loading
        worksheets = self.wb.get_sheet_names()
        for worksheet in worksheets:
            self.wb[worksheet].freeze_panes = self.wb[worksheet]['A2']

    def freeze_panes_first_row(self):
        self.ws.freeze_panes = self.ws['A2']

    def make_first_row_bold(self):
        boldStyle = Style(font=Font(bold=True))
        for row in self.ws[self.header_row_reference()]:
            for cell in row:
                cell.style = boldStyle

    def dict_reader(self, filename, tab_name, header_row_cell_value=''):
        """
        Creates list of ordered dictionaries of xlsx data (Recoded 4nov15)
        :param filename: xlsx filename to turn into dictionary
        :param tab_name: tab_name of the worksheet to read
        :param header_row_cell_value: the value of any cell indicating the header row
        :return: [{line1header1: value, ...},{line2header2: value}]
        """
        wb = load_workbook(filename=filename, read_only=True, data_only=True)
        ws = wb[tab_name]
        header = []
        result = []
        found = False if header_row_cell_value else True
        for r in ws.rows:
            if not found:
                for cell in r:
                    if cell.value == header_row_cell_value:
                        found = True
                        break
            if found:
                if not header:
                    header = [cell.value for cell in r]
                    continue
                result.append(OrderedDict(zip(header,
                                              [cell.value.strftime('%H:%M') if type(cell.value) is datetime.time
                                               else '00:00' if cell.value == datetime.datetime(1899, 12, 30, 0, 0)
                                               else cell.value.strftime('%d/%m/%Y') if type(cell.value) is datetime.datetime
                                               else '' if cell.value is None
                                               else str(cell.value)
                                               for cell in r])))
        return result

    def xlsx_to_csv(self, xlsx_to_read, csv_to_write, tab_name, delimeter=',', data_only=True, header_row_cell_value=''):
        # TODO: remove this method in re of using dict_reader + csv: dict_to_csv
        """
        Create csv from xlsx
        data_only=True: output value, not formula definition
        """
        wb = load_workbook(filename=xlsx_to_read, read_only=True, data_only=data_only)
        ws = wb[tab_name]
        with open(csv_to_write, 'w', newline='') as csvOut:
            cw = csv.writer(csvOut, delimiter=delimeter)
            found = False if header_row_cell_value else True
            for r in ws.rows:
                if not found:
                    for cell in r:
                        if cell.value == header_row_cell_value:
                            found = True
                            break
                if found:
                    cw.writerow([cell.value.strftime('%H:%M') if type(cell.value) is datetime.time
                                 else cell.value.strftime('%d/%m/%Y') if type(cell.value) is datetime.datetime
                                 else cell.value
                                 for cell in r])


        # data = self.dict_reader(xlsx_to_read, tab_name, header_row_cell_value=header_row_cell_value)
        # with open(csv_to_write, 'w', newline='') as csv_file:
        #     writer = csv.DictWriter(csv_file, list(data[0].keys()), delimiter=delimiter)
        #     writer.writeheader()
        #     for row in data:
        #         writer.writerow(row)